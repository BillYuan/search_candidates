import os
import logging

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


logger = logging.getLogger(u"sr")


class ExcelParser(object):
    """
    Excel Parser, use to read or write data from an excel file
    Data structure as below.
    ExcelParser:
        Sheet:
            ||HeaderA||HeaderB||
            | ValueA1 |ValueB1 |
            | ValueA2 |ValueB2 |
    Excel Data:
        [{  "sheetname": "sheetname1",
            "header": [HeaderA, HeaderB],
            "data": [{"HeaderA":"ValueA1", "HeaderB":"ValueB1"},{"HeaderA":"ValueA2", "HeaderB":"ValueB2"}]
         }],
         ...
        ]

    Features:
    1. Read mode, load all data from table to List-Dict, support read by special sheet name
    2. Write mode, write all data from List-Dict to table
    3. Bold for header
    """
    MODE_READ = "read"
    MODE_WRITE = "write"

    def __init__(self, path, mode=MODE_READ, excelData=None):
        self.path = path
        self.mode = mode

        if not excelData:
            self.excelData = []
        else:
            self.excelData = excelData
        self.sheets = []
        self.sheetsList = []

        if self.MODE_WRITE == self.mode:
            if os.path.exists(self.path):
                try:
                    os.remove(self.path)
                except Exception:
                    logger.exception(u"Cannot delete file {}, please check if it's opened!\n".format(self.path))
                    exit(1)

            self.wb = Workbook()
            if not self.excelData:
                return

            for sheetData in self.excelData:
                if not sheetData:
                    continue

                try:
                    if ExcelParser.Sheet.DATA not in sheetData or ExcelParser.Sheet.HEADER not in sheetData:
                        continue

                    sheetName = ExcelParser.Sheet.DEFAULT_NAME
                    if ExcelParser.Sheet.NAME in sheetData:
                        sheetName = sheetData[ExcelParser.Sheet.NAME]
                    if 0 == len(self.sheets):
                        ws = self.wb.active
                        ws.title = sheetName
                    else:
                        ws = self.wb.create_sheet(title=sheetName)

                    sheet = ExcelParser.Sheet(ws, headerKeys=sheetData[ExcelParser.Sheet.HEADER],
                                              rowData=sheetData[ExcelParser.Sheet.DATA])
                    self.__append_sheet(sheet, sheetName)
                except Exception:
                    logger.exception(u"Cannot create sheet by data: {}!\n".format(sheetData))
        else:
            # if not write mode, default is read mode
            self.wb = load_workbook(filename=self.path)
            sheetNames = self.wb.sheetnames
            for sheetName in sheetNames:
                sheet = self.__load_sheet(sheetName)
                self.__append_sheet(sheet, sheetName)

    def __load_sheet(self, name):
        ws = self.wb[name]
        return ExcelParser.Sheet(ws)

    def __append_sheet(self, sheet, sheetName):
        if sheetName not in self.sheetsList:
            self.sheets.append(sheet)
            self.sheetsList.append(sheetName)

    def __repr__(self):
        return (u"ExcelParser(path: {}, mode: {}, sheets len: {})"
                .format(self.path, self.mode, len(self.sheets)))

    def __len__(self):
        return len(self.sheets)

    def __getitem__(self, position):
        return self.sheets[position]

    def fetch(self):
        """ fetch all required sheets data
        """
        if not self.sheets:
            return None
        self.excelData = []
        for sheet in self.sheets:
            try:
                sheet.fetch()
                if sheet.header and sheet.data:
                    sheetInfo = {
                        ExcelParser.Sheet.NAME: sheet.get_name(),
                        ExcelParser.Sheet.HEADER: sheet.header,
                        ExcelParser.Sheet.DATA: sheet.data
                    }
                    self.excelData.append(sheetInfo)
            except Exception:
                logger.exception(u"Fetch sheet '{}' failed!\n".format(sheet))
        logger.debug(u"Sheets data: {}".format(self.excelData))
        return self.excelData

    def export(self):
        """ export all data to different sheet and save to a xlsx file
        """
        if not self.sheets:
            return None
        try:
            for sheet in self.sheets:
                sheet.export()
            # save file
            self.wb.save(filename=self.path)
            logger.info(u"Exported to excel: {}".format(self.path))
        except Exception:
            logger.exception(u"Export to excel {} failed!\n".format(self.path))

    class Sheet(object):
        """
        Excel Sheet
        """
        DEFAULT_NAME = "candidates"

        NAME = "sheetname"
        HEADER = "header"
        DATA = "data"

        MAX_COLUMN_SIZE = 120

        def __init__(self, worksheet, headerKeys=None, rowData=None):
            self.ws = worksheet

            if headerKeys:
                self.header = headerKeys
            else:
                self.header = []

            if rowData:
                self.data = rowData
            else:
                self.data = []

            logger.debug(u"header size: {}, data size: {}".format(len(self.header), len(self.data)))

        def __repr__(self):
            return "Sheet(sheetName: {}, headerKey len: {}, rowData len: {})" \
                .format(self.ws.title, len(self.header), len(self.data))

        def __len__(self):
            return len(self.data)

        def __getitem__(self, position):
            return self.data[position]

        def get_name(self):
            """ Get sheet name from work sheet
            Return none if no work sheet
            """
            if not self.ws:
                return None
            return self.ws.title

        def fetch(self):
            """ fetch
            fetch all data from excel

            :rtype: headerInfo, rowDataInfo
            """
            global rowIndex, colIndex
            if not self.ws:
                return

            try:
                # loop all data value
                rowIndex = 1
                for row in self.ws.values:
                    colIndex = 1
                    rowLine = {}
                    for cellValue in row:
                        if 1 == rowIndex:  # first line is key
                            if cellValue:
                                self.header.append(cellValue)
                            else:
                                logger.error(u"Excel first line header should include continual values,"
                                             u" please double check the excel table!")
                                exit(1)
                        else:
                            if cellValue:
                                rowLine[self.header[colIndex - 1]] = cellValue
                        colIndex += 1
                    if 1 != rowIndex:
                        self.data.append(rowLine)
                    rowIndex += 1
            except Exception:
                logger.exception(u"Fetch from excel sheet {} failed in row {}, col {} !\n\n"
                                 .format(self.ws.title, rowIndex, colIndex))
                self.data = []

            logger.debug(self.header)
            logger.debug(self.data)

        def export(self):
            """ export
            export all data to an excel
            """
            # pre defined font
            fontBold = Font(bold=True)

            # fill the header first
            columnMaxLen = {}
            for col, keyHeader in enumerate(self.header):
                if not keyHeader:
                    continue
                c = self.ws.cell(column=col + 1, row=1, value=keyHeader)
                c.font = fontBold
                # initialize the column length by key
                columnMaxLen[keyHeader] = len(keyHeader)

            # fill the content
            row = 2  # start from row 2
            for fields in self.data:
                for col, keyHeader in enumerate(self.header):
                    if not keyHeader:
                        continue
                    if keyHeader not in fields:
                        continue

                    cellValue = fields[keyHeader]

                    if not cellValue:
                        continue
                    logger.debug(u"fill (column={}, row={}), value={}".format(col + 1, row, cellValue))
                    encodeValue = cellValue

                    try:
                        c = self.ws.cell(column=col + 1, row=row, value=encodeValue)
                    except Exception as e:
                        logging.warning(u"Cannot fill cell to the excel at column {}, row {}!".format(col+1, row))
                        c = self.ws.cell(column=col + 1, row=row, value="Warning! Unable to fill excel Cell!")

                    # record the max column value length
                    if len(cellValue) > columnMaxLen[keyHeader]:
                        columnMaxLen[keyHeader] = len(cellValue)

                row += 1

            # adjust column size
            logger.debug(u"adjust width: {}".format(columnMaxLen))
            for col, keyHeader in enumerate(self.header):
                if not keyHeader:
                    continue
                adjusted_width = min((columnMaxLen[keyHeader]*2 + 1), self.MAX_COLUMN_SIZE)
                self.ws.column_dimensions[get_column_letter(col + 1)].width = adjusted_width
